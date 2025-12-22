# excel_export_local.py

from pathlib import Path
from datetime import datetime
from zoneinfo import ZoneInfo
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

EXPORT_DIR = Path(__file__).resolve().parents[2] / "exports"
EXPORT_DIR.mkdir(parents=True, exist_ok=True)

COLUMNS = ["Fecha","Pais","Campo","Pagina Web","Nivel","Remoto?","Empresa","Link","Salario","Aceptado?"]

EXCEL_PATH = EXPORT_DIR / "job_offers.xlsx"
SHEET_NAME = "Solicitudes_2025"

def date_madrid_str(fmt="%d %b %Y"):
    return datetime.now(ZoneInfo("Europe/Madrid")).strftime(fmt)

def _normalize_row(data: dict) -> dict:
    """Asegura claves exactas y orden; rellena vacíos; castea a str donde convenga."""
    row = {col: data.get(col, "") for col in COLUMNS}
    # defaults útiles
    if not row["Fecha"]:
        row["Fecha"] = date_madrid_str()
    # fuerza tipos de texto (evita NaN/None)
    for k in COLUMNS:
        if row[k] is None:
            row[k] = ""
    return row

def _add_or_update_table(file_path: Path, sheet_name: str):
    """
    Crea/expande una tabla con estilo que cubra A1:... toda la zona de datos.
    Compatible con distintas versiones de openpyxl (sin asignar ws._tables = []).
    """
    wb = load_workbook(file_path)
    ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active

    end_col = get_column_letter(len(COLUMNS))
    ref = f"A1:{end_col}{ws.max_row}"

    # 1) Intentar expandir una tabla existente (distintas representaciones)
    existing = None
    try:
        t = ws.tables
        # dict-like (openpyxl 3.x típico)
        if hasattr(t, "values") and not isinstance(t, (list, tuple, set)):
            vals = list(t.values())
            if vals and hasattr(vals[0], "ref"):
                existing = vals[0]
        # lista/iterable de objetos Table (algunas variantes)
        elif isinstance(t, (list, tuple, set)):
            for x in t:
                if hasattr(x, "ref"):
                    existing = x
                    break
    except Exception:
        existing = None

    if existing and hasattr(existing, "ref"):
        # Expandimos rango
        existing.ref = ref
    else:
        # 2) Eliminar tablas previas de forma segura (sin mutar el tipo interno)
        try:
            if hasattr(ws, "tables"):
                # dict-like con keys
                if hasattr(ws.tables, "keys"):
                    for name in list(ws.tables.keys()):
                        try:
                            del ws.tables[name]
                        except Exception:
                            pass
        except Exception:
            pass
        # Fallback: limpiar contenedor si tiene 'clear' (pero sin reasignar lista nueva)
        try:
            if hasattr(ws, "_tables") and hasattr(ws._tables, "clear"):
                ws._tables.clear()
        except Exception:
            pass

        # 3) Crear tabla nueva
        tbl = Table(displayName="Tabla_1", ref=ref)
        tbl.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False, showLastColumn=False,
            showRowStripes=True, showColumnStripes=False
        )
        try:
            ws.add_table(tbl)
        except Exception:
            # Fallback muy antiguo: añadir a la colección interna si es lista
            if hasattr(ws, "_tables") and hasattr(ws._tables, "append"):
                ws._tables.append(tbl)

    wb.save(file_path)



def export_to_excel(data: dict) -> Path:
    """Append de UNA fila al libro fijo, conservando formato/tabla."""
    row = _normalize_row(data)
    df = pd.DataFrame([row], columns=COLUMNS)

    if not EXCEL_PATH.exists():
        # crea el libro con headers
        with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl", mode="w") as w:
            df.to_excel(w, index=False, sheet_name=SHEET_NAME)
        _add_or_update_table(EXCEL_PATH, SHEET_NAME)
        return EXCEL_PATH

    # calcula la fila de comienzo sin tocar writer.book
    wb = load_workbook(EXCEL_PATH)
    ws = wb[SHEET_NAME] if SHEET_NAME in wb.sheetnames else wb.active
    startrow = ws.max_row + 1
    wb.close()

    # escribe sin encabezado a partir de startrow (overlay)
    with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl", mode="a", if_sheet_exists="overlay") as w:
        df.to_excel(w, index=False, header=False, sheet_name=SHEET_NAME, startrow=startrow-1)

    # ajusta/crea la tabla tras escribir
    _add_or_update_table(EXCEL_PATH, SHEET_NAME)
    return EXCEL_PATH
